"""
Excel Export Module

Exports extracted data, classified documents, matched fields, and semantic matching results to Excel.
Creates formatted, well-organized Excel workbooks with multiple sheets.
"""

import os
from pathlib import Path
import csv


def _try_import_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter
    except Exception:
        return None, None, None, None, None, None, None


class ExcelExporter:
    """
    Exports document processing results to Excel workbooks.
    Handles multiple sheets with formatting and headers.
    """
    
    def __init__(self):
        """Initialize style templates."""
        openpyxl_mod, Font, PatternFill, Alignment, Border, Side, get_column_letter = _try_import_openpyxl()

        if openpyxl_mod:
            self._has_openpyxl = True
            self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            self.header_font = Font(bold=True, color='FFFFFF', size=11)
            self.title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            self.title_font = Font(bold=True, color='FFFFFF', size=12)
            self.matched_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            self.unmatched_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            
            self.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            self.alignment_wrap = Alignment(wrap_text=True, vertical='top')
        else:
            self._has_openpyxl = False
            # fallback fills/fonts to None when openpyxl isn't present
            self.header_fill = None
            self.header_font = None
            self.title_fill = None
            self.title_font = None
            self.matched_fill = None
            self.unmatched_fill = None
            self.border = None
            self.alignment_wrap = None
    
    def _set_cell_style(self, cell, bold=False, fill=None, font_color='000000'):
        """Apply styling to a cell."""
        # No-op when openpyxl not available
        if not self._has_openpyxl:
            return
        cell.border = self.border
        cell.alignment = self.alignment_wrap
        if bold:
            from openpyxl.styles import Font as _Font
            cell.font = _Font(bold=True, color=font_color)
        if fill:
            cell.fill = fill
    
    def export_extraction_results(self, extracted_data: dict, output_path: str):
        """
        Export extracted fields from documents to Excel.
        
        Args:
            extracted_data (dict): Fields extracted from each document
            output_path (str): Path to output Excel file
        """
        
        openpyxl_mod, *_ = _try_import_openpyxl()
        if openpyxl_mod is None:
            # Fallback: write CSV instead
            csv_path = os.path.splitext(output_path)[0] + '.csv'
            os.makedirs(os.path.dirname(csv_path), exist_ok=True)
            headers = ['Document Name', 'Field', 'Value', 'Confidence (%)', 'Extraction Method', 'Notes']
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                for doc_name, fields in extracted_data.items():
                    for field_name, field_data in fields.items():
                        field_display = field_name.replace('_', ' ').title()
                        value = field_data.get('name') or field_data.get('organization') or field_data.get('currency') or field_data.get('address') or 'Not Found'
                        confidence = field_data.get('confidence', 0)
                        method = field_data.get('method', 'unknown')
                        explanation = field_data.get('explanation', '')
                        writer.writerow([doc_name, field_display, value, confidence, method, explanation])
            print(f"openpyxl not available; wrote CSV to: {csv_path}\n")
            return
        wb = openpyxl_mod.Workbook()
        ws = wb.active
        ws.title = "Extracted Fields"
        
        # Create headers
        headers = ['Document Name', 'Field', 'Value', 'Confidence (%)', 'Extraction Method', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._set_cell_style(cell, bold=True, fill=self.header_fill, font_color='FFFFFF')
        
        # Add data
        row = 2
        for doc_name, fields in extracted_data.items():
            for field_name, field_data in fields.items():
                
                # Get field display name
                field_display = field_name.replace('_', ' ').title()
                
                # Determine value and confidence
                value = None
                confidence = field_data.get('confidence', 0)
                method = field_data.get('method', 'unknown')
                explanation = field_data.get('explanation', '')
                
                if 'name' in field_data:
                    value = field_data['name']
                elif 'organization' in field_data:
                    value = field_data['organization']
                elif 'currency' in field_data:
                    value = field_data['currency']
                elif 'address' in field_data:
                    value = field_data['address']
                
                # Write row
                ws.cell(row=row, column=1, value=doc_name)
                ws.cell(row=row, column=2, value=field_display)
                ws.cell(row=row, column=3, value=value if value else 'Not Found')
                ws.cell(row=row, column=4, value=confidence)
                ws.cell(row=row, column=5, value=method)
                ws.cell(row=row, column=6, value=explanation)
                
                # Highlight not found rows
                if not value:
                    fill_color = self.unmatched_fill
                else:
                    fill_color = None
                
                for col in range(1, 7):
                    self._set_cell_style(ws.cell(row=row, column=col), fill=fill_color)
                
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 35
        
        # Save workbook
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"Extraction results exported to: {output_path}\n")
    
    def export_classification_results(self, classifications: dict, output_path: str):
        """
        Export document classification results to Excel.
        
        Args:
            classifications (dict): Classification results for each document
            output_path (str): Path to output Excel file
        """
        
        openpyxl_mod, *_ = _try_import_openpyxl()
        if openpyxl_mod is None:
            # Fallback CSV
            csv_path = os.path.splitext(output_path)[0] + '.csv'
            os.makedirs(os.path.dirname(csv_path), exist_ok=True)
            headers = ['Document Name', 'Classified Type', 'Confidence (%)', 'Keywords Found', 'Match Counts']
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                for doc_name, classification in classifications.items():
                    doc_type = classification.get('document_type', 'Unknown')
                    confidence = classification.get('confidence', 0)
                    keywords = ', '.join(classification.get('found_keywords', []))
                    match_counts = classification.get('match_counts', {})
                    counts_str = f"Q:{match_counts.get('quotation', 0)}, SOW:{match_counts.get('sow', 0)}, C:{match_counts.get('contract', 0)}"
                    writer.writerow([doc_name, doc_type, confidence, keywords, counts_str])
            print(f"openpyxl not available; wrote CSV to: {csv_path}\n")
            return
        wb = openpyxl_mod.Workbook()
        ws = wb.active
        ws.title = "Document Classification"
        
        # Create headers
        headers = ['Document Name', 'Classified Type', 'Confidence (%)', 'Keywords Found', 'Match Counts']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._set_cell_style(cell, bold=True, fill=self.header_fill, font_color='FFFFFF')
        
        # Add data
        row = 2
        for doc_name, classification in classifications.items():
            doc_type = classification.get('document_type', 'Unknown')
            confidence = classification.get('confidence', 0)
            keywords = ', '.join(classification.get('found_keywords', []))
            match_counts = classification.get('match_counts', {})
            counts_str = f"Q:{match_counts.get('quotation', 0)}, SOW:{match_counts.get('sow', 0)}, C:{match_counts.get('contract', 0)}"
            
            ws.cell(row=row, column=1, value=doc_name)
            ws.cell(row=row, column=2, value=doc_type)
            ws.cell(row=row, column=3, value=confidence)
            ws.cell(row=row, column=4, value=keywords)
            ws.cell(row=row, column=5, value=counts_str)
            
            for col in range(1, 6):
                self._set_cell_style(ws.cell(row=row, column=col))
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 25
        
        # Save workbook
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"Classification results exported to: {output_path}\n")
    
    def export_matching_results(self, matching_data: list, output_path: str):
        """
        Export semantic matching results to Excel.
        
        Args:
            matching_data (list): List of matching results from all document pairs
            output_path (str): Path to output Excel file
        """
        
        openpyxl_mod, *_ = _try_import_openpyxl()
        if openpyxl_mod is None:
            # Fallback CSV for matching data
            csv_path = os.path.splitext(output_path)[0] + '.csv'
            os.makedirs(os.path.dirname(csv_path), exist_ok=True)
            headers = ['Document Pair', 'Total Fields', 'Matched', 'Likely Matched', 'Overall Score (/100)']
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                for match_result in matching_data:
                    doc_pair = f"{match_result['document_1']} <-> {match_result['document_2']}"
                    total = match_result['total_fields_compared']
                    matched = match_result['matched_fields']
                    likely = match_result['likely_matched_fields']
                    score = match_result['overall_match_score']
                    writer.writerow([doc_pair, total, matched, likely, score])
            print(f"openpyxl not available; wrote CSV to: {csv_path}\n")
            return
        wb = openpyxl_mod.Workbook()
        
        # Create summary sheet
        ws_summary = wb.active
        ws_summary.title = "Matching Summary"
        
        # Summary headers
        headers = ['Document Pair', 'Total Fields', 'Matched', 'Likely Matched', 'Overall Score (/100)']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            self._set_cell_style(cell, bold=True, fill=self.header_fill, font_color='FFFFFF')
        
        # Summary data
        row = 2
        for match_result in matching_data:
            doc_pair = f"{match_result['document_1']} <-> {match_result['document_2']}"
            total = match_result['total_fields_compared']
            matched = match_result['matched_fields']
            likely = match_result['likely_matched_fields']
            score = match_result['overall_match_score']
            
            ws_summary.cell(row=row, column=1, value=doc_pair)
            ws_summary.cell(row=row, column=2, value=total)
            ws_summary.cell(row=row, column=3, value=matched)
            ws_summary.cell(row=row, column=4, value=likely)
            ws_summary.cell(row=row, column=5, value=score)
            
            for col in range(1, 6):
                self._set_cell_style(ws_summary.cell(row=row, column=col))
            
            row += 1
        
        # Adjust column widths for summary
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 14
        ws_summary.column_dimensions['C'].width = 12
        ws_summary.column_dimensions['D'].width = 16
        ws_summary.column_dimensions['E'].width = 18
        
        # Create detailed sheet for each match pair
        for idx, match_result in enumerate(matching_data, 1):
            ws = wb.create_sheet(title=f"Match Pair {idx}")
            
            # Title
            doc_pair = f"{match_result['document_1']} <-> {match_result['document_2']}"
            title_cell = ws.cell(row=1, column=1, value=f"Detailed Matching: {doc_pair}")
            title_cell.font = self.title_font
            title_cell.fill = self.title_fill
            ws.merge_cells('A1:F1')
            
            # Headers
            headers = ['Field', 'Status', 'Value 1', 'Confidence 1 (%)', 'Value 2', 'Confidence 2 (%)', 'Match Score (/100)', 'Explanation']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                self._set_cell_style(cell, bold=True, fill=self.header_fill, font_color='FFFFFF')
            
            # Data
            row = 4
            for field_match in match_result['matching_results']:
                status = field_match['status']
                
                # Determine fill based on match status
                if 'Matched' in status:
                    fill_color = self.matched_fill
                else:
                    fill_color = self.unmatched_fill
                
                ws.cell(row=row, column=1, value=field_match['field'])
                ws.cell(row=row, column=2, value=status)
                ws.cell(row=row, column=3, value=field_match['value1'] if field_match['value1'] else 'Not Found')
                ws.cell(row=row, column=4, value=field_match['confidence_1'])
                ws.cell(row=row, column=5, value=field_match['value2'] if field_match['value2'] else 'Not Found')
                ws.cell(row=row, column=6, value=field_match['confidence_2'])
                ws.cell(row=row, column=7, value=field_match['match_score'])
                ws.cell(row=row, column=8, value=field_match['explanation'])
                
                for col in range(1, 9):
                    self._set_cell_style(ws.cell(row=row, column=col), fill=fill_color)
                
                row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 16
            ws.column_dimensions['H'].width = 40
        
        # Save workbook
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"Matching results exported to: {output_path}\n")
    
    def export_complete_report(self, extracted_data: dict, classifications: dict, 
                             matching_data: list, output_path: str):
        """
        Export a comprehensive report with all results.
        
        Args:
            extracted_data (dict): Extracted fields
            classifications (dict): Document classifications
            matching_data (list): Semantic matching results
            output_path (str): Path to output Excel file
        """
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # 1. Summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary.cell(row=1, column=1, value="INTELLIGENT DOCUMENT AUTOMATION - FINAL REPORT").font = self.title_font
        ws_summary.cell(row=1, column=1).fill = self.title_fill
        ws_summary.merge_cells('A1:D1')
        
        ws_summary.cell(row=3, column=1, value="Processing Summary:")
        ws_summary.cell(row=4, column=1, value=f"Total Documents: {len(classifications)}")
        ws_summary.cell(row=5, column=1, value=f"Document Pairs Matched: {len(matching_data)}")
        
        if matching_data:
            avg_score = sum(m['overall_match_score'] for m in matching_data) / len(matching_data)
            ws_summary.cell(row=6, column=1, value=f"Average Match Score: {avg_score:.2f}/100")
        
        # 2. Classification sheet
        ws_class = wb.create_sheet("Document Classification")
        headers = ['Document Name', 'Classified Type', 'Confidence (%)', 'Keywords Found']
        for col, header in enumerate(headers, 1):
            cell = ws_class.cell(row=1, column=col, value=header)
            self._set_cell_style(cell, bold=True, fill=self.header_fill, font_color='FFFFFF')
        
        row = 2
        for doc_name, classification in classifications.items():
            ws_class.cell(row=row, column=1, value=doc_name)
            ws_class.cell(row=row, column=2, value=classification.get('document_type', 'Unknown'))
            ws_class.cell(row=row, column=3, value=classification.get('confidence', 0))
            ws_class.cell(row=row, column=4, value=', '.join(classification.get('found_keywords', [])))
            
            for col in range(1, 5):
                self._set_cell_style(ws_class.cell(row=row, column=col))
            row += 1
        
        ws_class.column_dimensions['A'].width = 25
        ws_class.column_dimensions['B'].width = 20
        ws_class.column_dimensions['C'].width = 14
        ws_class.column_dimensions['D'].width = 35
        
        # 3. Extracted Fields sheet
        ws_extract = wb.create_sheet("Extracted Fields")
        headers = ['Document Name', 'Field', 'Value', 'Confidence (%)', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws_extract.cell(row=1, column=col, value=header)
            self._set_cell_style(cell, bold=True, fill=self.header_fill, font_color='FFFFFF')
        
        row = 2
        for doc_name, fields in extracted_data.items():
            for field_name, field_data in fields.items():
                field_display = field_name.replace('_', ' ').title()
                
                value = None
                if 'name' in field_data:
                    value = field_data['name']
                elif 'organization' in field_data:
                    value = field_data['organization']
                elif 'currency' in field_data:
                    value = field_data['currency']
                elif 'address' in field_data:
                    value = field_data['address']
                
                ws_extract.cell(row=row, column=1, value=doc_name)
                ws_extract.cell(row=row, column=2, value=field_display)
                ws_extract.cell(row=row, column=3, value=value if value else 'Not Found')
                ws_extract.cell(row=row, column=4, value=field_data.get('confidence', 0))
                ws_extract.cell(row=row, column=5, value=field_data.get('explanation', ''))
                
                fill_color = None if value else self.unmatched_fill
                for col in range(1, 6):
                    self._set_cell_style(ws_extract.cell(row=row, column=col), fill=fill_color)
                row += 1
        
        ws_extract.column_dimensions['A'].width = 22
        ws_extract.column_dimensions['B'].width = 18
        ws_extract.column_dimensions['C'].width = 25
        ws_extract.column_dimensions['D'].width = 14
        ws_extract.column_dimensions['E'].width = 35
        
        # 4. Matching Results sheet
        if matching_data:
            ws_match = wb.create_sheet("Semantic Matching")
            headers = ['Document Pair', 'Field', 'Status', 'Value 1', 'Value 2', 'Match Score (/100)', 'Explanation']
            for col, header in enumerate(headers, 1):
                cell = ws_match.cell(row=1, column=col, value=header)
                self._set_cell_style(cell, bold=True, fill=self.header_fill, font_color='FFFFFF')
            
            row = 2
            for match_result in matching_data:
                doc_pair = f"{match_result['document_1']} <-> {match_result['document_2']}"
                
                for field_match in match_result['matching_results']:
                    status = field_match['status']
                    fill_color = self.matched_fill if 'Matched' in status else self.unmatched_fill
                    
                    ws_match.cell(row=row, column=1, value=doc_pair)
                    ws_match.cell(row=row, column=2, value=field_match['field'])
                    ws_match.cell(row=row, column=3, value=status)
                    ws_match.cell(row=row, column=4, value=field_match['value1'] if field_match['value1'] else 'Not Found')
                    ws_match.cell(row=row, column=5, value=field_match['value2'] if field_match['value2'] else 'Not Found')
                    ws_match.cell(row=row, column=6, value=field_match['match_score'])
                    ws_match.cell(row=row, column=7, value=field_match['explanation'])
                    
                    for col in range(1, 8):
                        self._set_cell_style(ws_match.cell(row=row, column=col), fill=fill_color)
                    row += 1
            
            ws_match.column_dimensions['A'].width = 30
            ws_match.column_dimensions['B'].width = 18
            ws_match.column_dimensions['C'].width = 15
            ws_match.column_dimensions['D'].width = 20
            ws_match.column_dimensions['E'].width = 20
            ws_match.column_dimensions['F'].width = 16
            ws_match.column_dimensions['G'].width = 40
        
        # Save workbook
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"Complete report exported to: {output_path}\n")
    
    def export_hybrid_matching_results(self, matching_results: dict, output_path: str):
        """
        Export hybrid pattern-semantic matching results to Excel.
        Includes detailed layer-by-layer analysis with numeric scores.
        
        Args:
            matching_results (dict): Results from HybridMatchingEngine
            output_path (str): Path to output Excel file
        """
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        title_cell = ws_summary.cell(row=1, column=1, value="HYBRID PATTERN-SEMANTIC MATCHING RESULTS")
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        ws_summary.merge_cells('A1:H1')
        
        ws_summary.cell(row=3, column=1, value="Framework Components:")
        ws_summary.cell(row=4, column=1, value="✓ Pattern Matching Layer (difflib SequenceMatcher)")
        ws_summary.cell(row=5, column=1, value="✓ Rule-Based Validation Layer (Currency normalization)")
        ws_summary.cell(row=6, column=1, value="✓ Semantic Similarity Layer (Token-based Jaccard, LLM-inspired)")
        ws_summary.cell(row=7, column=1, value="✓ Decision Fusion Engine (0.6×Pattern + 0.4×Semantic)")
        ws_summary.cell(row=8, column=1, value="✓ Explainable Output (All scores shown transparently)")
        
        ws_summary.merge_cells('A4:C4')
        ws_summary.merge_cells('A5:C5')
        ws_summary.merge_cells('A6:C6')
        ws_summary.merge_cells('A7:C7')
        ws_summary.merge_cells('A8:C8')
        
        # Details - create a sheet for each document pair
        details = matching_results.get('details', {})
        
        for pair_idx, (pair_name, fields_results) in enumerate(details.items(), 1):
            
            ws = wb.create_sheet(title=f"Pair {pair_idx}")
            
            # Title with document pair names
            title = f"Detailed Analysis: {pair_name}"
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = self.title_font
            title_cell.fill = self.title_fill
            ws.merge_cells(f'A1:F1')
            
            # Headers for results table
            headers = [
                'Field',
                'Status',
                'Pattern Score (0-1)',
                'Semantic Score (0-1)',
                'Final Score (0-1)',
                'Decision'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                self._set_cell_style(cell, bold=True, fill=self.header_fill, font_color='FFFFFF')
            
            # Add field results
            row = 4
            for field_name, field_result in fields_results.items():
                
                status = field_result.get('status', 'NOT FOUND')
                final_score = field_result.get('final_score', 0)
                pattern_score = field_result.get('pattern_score', 0)
                semantic_score = field_result.get('semantic_score', 0)
                
                # Color code based on score
                if final_score >= 0.75:
                    fill_color = self.matched_fill
                    decision = "MATCH ✓"
                else:
                    fill_color = self.unmatched_fill
                    decision = "NO MATCH ✗"
                
                ws.cell(row=row, column=1, value=field_name)
                ws.cell(row=row, column=2, value=status)
                ws.cell(row=row, column=3, value=f"{pattern_score:.2f}")
                ws.cell(row=row, column=4, value=f"{semantic_score:.2f}")
                ws.cell(row=row, column=5, value=f"{final_score:.2f}")
                ws.cell(row=row, column=6, value=decision)
                
                for col in range(1, 7):
                    self._set_cell_style(ws.cell(row=row, column=col), fill=fill_color)
                
                row += 1
            
            # Add explanation section
            row += 2
            ws.cell(row=row, column=1, value="Explanation Details").font = Font(bold=True, size=11)
            row += 1
            
            for field_name, field_result in fields_results.items():
                explanation_lines = field_result.get('explanation', [])
                
                ws.cell(row=row, column=1, value=f"Field: {field_name}").font = Font(bold=True)
                row += 1
                
                for line in explanation_lines:
                    if line.strip():
                        ws.cell(row=row, column=1, value=line)
                        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
                    row += 1
                
                row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
        
        # Save
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"Hybrid matching results exported to: {output_path}\n")
